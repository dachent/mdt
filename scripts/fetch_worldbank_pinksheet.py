#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen


HISTORICAL_MONTHLY_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/CMO-Historical-Data-Monthly.xlsx"
)
HISTORICAL_ANNUAL_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/CMO-Historical-Data-Annual.xlsx"
)
USER_AGENT = "Codex agent-market-data-terminal/1.0"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch the World Bank Pink Sheet (CMO) commodity-price workbook.")
    parser.add_argument(
        "--frequency",
        choices=["monthly", "annual"],
        default="monthly",
        help="Workbook variant. Defaults to monthly.",
    )
    parser.add_argument(
        "--url",
        help="Optional override URL when the CMO release identifier rotates.",
    )
    parser.add_argument("--sheet", help="Optional sheet name override.")
    parser.add_argument("--format", required=True, choices=["raw", "parsed"])
    parser.add_argument("--output", required=True, help="Output path.")
    return parser.parse_args()


def resolve_url(args: argparse.Namespace) -> str:
    if args.url:
        return args.url
    if args.frequency == "annual":
        return HISTORICAL_ANNUAL_URL
    return HISTORICAL_MONTHLY_URL


def fetch_bytes(url: str) -> bytes:
    request = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(request, timeout=120) as response:
        return response.read()


def select_sheet_name(sheet_names: list[str], override: str | None, frequency: str) -> str:
    if override:
        if override in sheet_names:
            return override
        raise RuntimeError(f"--sheet {override!r} not found in workbook. Sheets: {sheet_names}")
    keyword = "monthly" if frequency == "monthly" else "annual"
    primary = [name for name in sheet_names if keyword in name.lower() and "price" in name.lower()]
    if primary:
        return primary[0]
    fallback = [name for name in sheet_names if keyword in name.lower()]
    if fallback:
        return fallback[0]
    return sheet_names[0]


def parse_workbook(payload: bytes, source_url: str, frequency: str, override_sheet: str | None) -> dict[str, object]:
    try:
        from openpyxl import load_workbook  # type: ignore
        from io import BytesIO
    except ImportError as exc:
        raise RuntimeError(
            "Parsed Pink Sheet mode requires openpyxl. Install with `pip install openpyxl`."
        ) from exc

    workbook = load_workbook(filename=BytesIO(payload), data_only=True, read_only=True)
    sheet_names = list(workbook.sheetnames)
    selected = select_sheet_name(sheet_names, override_sheet, frequency)
    sheet = workbook[selected]

    rows: list[list[object]] = []
    for row in sheet.iter_rows(values_only=True):
        if row is None:
            continue
        rows.append([cell for cell in row])

    header_index: int | None = None
    for index in range(min(len(rows), 12)):
        non_empty = [cell for cell in rows[index] if cell not in (None, "")]
        if len(non_empty) >= 3:
            header_index = index
            break
    if header_index is None:
        raise RuntimeError("Could not locate a header row in the selected sheet.")

    secondary_index = header_index + 1 if header_index + 1 < len(rows) else None
    primary_header = [
        str(cell) if cell is not None else f"column_{idx + 1}"
        for idx, cell in enumerate(rows[header_index])
    ]
    secondary_header: list[str] | None = None
    if secondary_index is not None:
        secondary_row = rows[secondary_index]
        secondary_header = [str(cell) if cell is not None else "" for cell in secondary_row]
        is_data_row = any(
            isinstance(cell, (int, float)) for cell in secondary_row if cell not in (None, "")
        )
        if is_data_row:
            secondary_header = None

    data_start = header_index + 1 if secondary_header is None else header_index + 2
    data_rows: list[dict[str, object]] = []
    for row in rows[data_start:]:
        if not row or all(cell in (None, "") for cell in row):
            continue
        record: dict[str, object] = {}
        for idx, column_name in enumerate(primary_header):
            if idx >= len(row):
                continue
            record[column_name] = row[idx]
        data_rows.append(record)

    return {
        "provider": "worldbank_pinksheet",
        "source_url": source_url,
        "frequency": frequency,
        "sheet_names": sheet_names,
        "selected_sheet": selected,
        "header": primary_header,
        "secondary_header": secondary_header,
        "rows": data_rows,
        "row_count": len(data_rows),
    }


def main() -> int:
    args = parse_args()
    output_path = Path(args.output).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    source_url = resolve_url(args)

    try:
        payload = fetch_bytes(source_url)
        if args.format == "raw":
            output_path.write_bytes(payload)
            print(f"Saved raw Pink Sheet payload to {output_path}")
            print(f"URL: {source_url}")
            return 0

        artifact = parse_workbook(payload, source_url, args.frequency, args.sheet)
        output_path.write_text(json.dumps(artifact, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
        print(f"Saved parsed Pink Sheet artifact to {output_path}")
        print(f"sheet={artifact['selected_sheet']} rows={artifact['row_count']}")
        return 0
    except (HTTPError, URLError, OSError, RuntimeError, ValueError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
