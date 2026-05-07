#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import html
import json
import re
import sys
from datetime import date
from io import BytesIO
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen


USER_AGENT = "Codex agent-market-data-terminal/1.0"

BUNDESBANK_SDMX_BASE = "https://api.statistiken.bundesbank.de/rest/data"
BOE_IADB_BASE = "https://www.bankofengland.co.uk/boeapps/database/_iadb-fromshowcolumns.asp"
RBA_F2_URL = "https://www.rba.gov.au/statistics/tables/xls/f02hist.xlsx"
BOC_VALET_BASE = "https://www.bankofcanada.ca/valet/observations"
BOJ_SEARCH_URL = "https://www.stat-search.boj.or.jp/ssi/cgi-bin/famecgi2?cgi=$nme_a000_en"

TAG_PATTERN = re.compile(r"<[^>]+>")
ROW_PATTERN = re.compile(r"<tr[^>]*>(.*?)</tr>", re.IGNORECASE | re.DOTALL)
CELL_PATTERN = re.compile(r"<t[dh][^>]*>(.*?)</t[dh]>", re.IGNORECASE | re.DOTALL)
TABLE_PATTERN = re.compile(r"<table\b[^>]*>(.*?)</table>", re.IGNORECASE | re.DOTALL)
MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch foreign central-bank historical bond yield archives.")
    parser.add_argument("--source", required=True, choices=["bundesbank", "boe", "rba", "boc", "boj"])
    parser.add_argument("--series", help="Series code or comma-separated codes (BoE, BoC, BoJ).")
    parser.add_argument("--key", help="Bundesbank series key (used as `{key}` in the SDMX URL).")
    parser.add_argument("--dataflow", default="BBK01", help="Bundesbank dataflow id. Defaults to BBK01.")
    parser.add_argument("--bundesbank-format", choices=["sdmx-json", "csv"], default="sdmx-json")
    parser.add_argument("--start", help="Start date in YYYY-MM-DD.")
    parser.add_argument("--end", help="End date in YYYY-MM-DD.")
    parser.add_argument("--sheet", help="RBA: override sheet name.")
    parser.add_argument("--url", help="Optional explicit override URL.")
    parser.add_argument("--format", required=True, choices=["raw", "parsed"])
    parser.add_argument("--output", required=True, help="Output path.")
    return parser.parse_args()


def fetch_bytes(url: str) -> bytes:
    request = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(request, timeout=120) as response:
        return response.read()


def clean_text(value: str) -> str:
    text = TAG_PATTERN.sub(" ", value)
    text = html.unescape(text).replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def parse_number(value: str) -> float | int | None:
    cleaned = clean_text(value).replace(",", "")
    if not cleaned or cleaned in {"-", "--", "n/a", "N/A"}:
        return None
    try:
        number = float(cleaned)
    except ValueError:
        return None
    if number.is_integer():
        return int(number)
    return number


def iso_to_boe_date(iso_value: str) -> str:
    parsed = date.fromisoformat(iso_value)
    return f"{parsed.day:02d}{MONTH_ABBR[parsed.month - 1]}{parsed.year}"


def build_bundesbank_url(args: argparse.Namespace) -> str:
    if args.url:
        return args.url
    if not args.key:
        raise ValueError("--key is required for source=bundesbank.")
    base = f"{BUNDESBANK_SDMX_BASE}/{args.dataflow}/{args.key}"
    params: list[tuple[str, str]] = []
    if args.bundesbank_format == "csv":
        params.append(("format", "csv"))
    if args.start:
        params.append(("startPeriod", args.start))
    if args.end:
        params.append(("endPeriod", args.end))
    return base + (f"?{urlencode(params)}" if params else "")


def build_boe_url(args: argparse.Namespace) -> str:
    if args.url:
        return args.url
    if not args.series:
        raise ValueError("--series is required for source=boe.")
    params: list[tuple[str, str]] = [("CSVF", "TT"), ("SeriesCodes", args.series), ("UsingCodes", "Y")]
    if args.start:
        params.append(("Datefrom", iso_to_boe_date(args.start)))
    if args.end:
        params.append(("Dateto", iso_to_boe_date(args.end)))
    return f"{BOE_IADB_BASE}?{urlencode(params)}"


def build_boc_url(args: argparse.Namespace) -> str:
    if args.url:
        return args.url
    if not args.series:
        raise ValueError("--series is required for source=boc.")
    fmt = "json" if args.format == "parsed" else "csv"
    params: list[tuple[str, str]] = []
    if args.start:
        params.append(("start_date", args.start))
    if args.end:
        params.append(("end_date", args.end))
    suffix = f"?{urlencode(params)}" if params else ""
    return f"{BOC_VALET_BASE}/{args.series}/{fmt}{suffix}"


def parse_csv_text(text: str) -> tuple[list[str], list[dict[str, object]]]:
    reader = csv.reader(text.splitlines())
    all_rows = [row for row in reader if row]
    if not all_rows:
        return [], []
    header = [cell.strip() for cell in all_rows[0]]
    data_rows: list[dict[str, object]] = []
    for cells in all_rows[1:]:
        record: dict[str, object] = {}
        for index, column_name in enumerate(header):
            if index >= len(cells):
                continue
            key = column_name or f"column_{index + 1}"
            value = cells[index]
            parsed = parse_number(value)
            record[key] = parsed if parsed is not None else value.strip()
        if record:
            data_rows.append(record)
    return header, data_rows


def parse_bundesbank_sdmx(text: str) -> dict[str, object]:
    payload = json.loads(text)
    data = payload.get("data") if isinstance(payload, dict) else None
    if not isinstance(data, dict):
        return {"raw_payload": payload}
    data_sets = data.get("dataSets", [])
    structures = data.get("structures", [])
    if not data_sets or not structures:
        return {"raw_payload": payload}
    structure = structures[0] if isinstance(structures, list) else {}
    dimensions = structure.get("dimensions", {}) if isinstance(structure, dict) else {}
    obs_dims = dimensions.get("observation", []) if isinstance(dimensions, dict) else []
    series_dims = dimensions.get("series", []) if isinstance(dimensions, dict) else []
    series_map = data_sets[0].get("series", {}) if isinstance(data_sets[0], dict) else {}

    rows: list[dict[str, object]] = []
    for series_key, series_payload in series_map.items() if isinstance(series_map, dict) else []:
        if not isinstance(series_payload, dict):
            continue
        series_indexes = series_key.split(":")
        base: dict[str, object] = {}
        for dim, idx in zip(series_dims, series_indexes):
            if not isinstance(dim, dict):
                continue
            try:
                position = int(idx)
            except ValueError:
                continue
            values = dim.get("values", [])
            if isinstance(values, list) and 0 <= position < len(values) and isinstance(values[position], dict):
                base[str(dim.get("id"))] = values[position].get("id")

        observations = series_payload.get("observations", {})
        if not isinstance(observations, dict):
            continue
        for obs_index_key, obs_payload in observations.items():
            row = dict(base)
            obs_indexes = obs_index_key.split(":")
            for dim, idx in zip(obs_dims, obs_indexes):
                if not isinstance(dim, dict):
                    continue
                try:
                    position = int(idx)
                except ValueError:
                    continue
                values = dim.get("values", [])
                if isinstance(values, list) and 0 <= position < len(values) and isinstance(values[position], dict):
                    row[str(dim.get("id"))] = values[position].get("id")
            if isinstance(obs_payload, list) and obs_payload:
                row["value"] = obs_payload[0]
            else:
                row["value"] = obs_payload
            rows.append(row)
    return {"rows": rows, "row_count": len(rows)}


def parse_html_tables(text: str) -> list[dict[str, object]]:
    tables: list[dict[str, object]] = []
    for table_match in TABLE_PATTERN.finditer(text):
        rows: list[list[str]] = []
        for row_html in ROW_PATTERN.findall(table_match.group(1)):
            cells = [clean_text(cell) for cell in CELL_PATTERN.findall(row_html)]
            if cells:
                rows.append(cells)
        if not rows:
            continue
        header = rows[0]
        data_rows: list[dict[str, object]] = []
        for cells in rows[1:]:
            record: dict[str, object] = {}
            for index, column_name in enumerate(header):
                if index >= len(cells):
                    continue
                key = column_name or f"column_{index + 1}"
                value = cells[index]
                parsed = parse_number(value)
                record[key] = parsed if parsed is not None else value
            if record:
                data_rows.append(record)
        tables.append({"header": header, "rows": data_rows, "row_count": len(data_rows)})
    return tables


def parse_rba_xlsx(payload: bytes, override_sheet: str | None) -> dict[str, object]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError("RBA parsing requires openpyxl. Install with `pip install openpyxl`.") from exc
    workbook = load_workbook(filename=BytesIO(payload), data_only=True, read_only=True)
    sheet_names = list(workbook.sheetnames)
    selected = override_sheet
    if selected is None:
        candidates = [name for name in sheet_names if "yield" in name.lower() or "f02" in name.lower()]
        selected = candidates[0] if candidates else sheet_names[0]
    if selected not in sheet_names:
        raise RuntimeError(f"--sheet {selected!r} not found. Sheets: {sheet_names}")
    sheet = workbook[selected]
    rows: list[list[object]] = []
    for row in sheet.iter_rows(values_only=True):
        if row is None:
            continue
        rows.append(list(row))

    header_index: int | None = None
    for index in range(min(len(rows), 12)):
        non_empty = [cell for cell in rows[index] if cell not in (None, "")]
        if len(non_empty) >= 2:
            header_index = index
            break
    if header_index is None:
        return {"sheet_names": sheet_names, "selected_sheet": selected, "rows": [], "row_count": 0}

    header = [
        str(cell) if cell is not None else f"column_{idx + 1}"
        for idx, cell in enumerate(rows[header_index])
    ]
    data_rows: list[dict[str, object]] = []
    for row in rows[header_index + 1 :]:
        if not row or all(cell in (None, "") for cell in row):
            continue
        record: dict[str, object] = {}
        for idx, column_name in enumerate(header):
            if idx >= len(row):
                continue
            record[column_name] = row[idx]
        data_rows.append(record)
    return {
        "sheet_names": sheet_names,
        "selected_sheet": selected,
        "header": header,
        "rows": data_rows,
        "row_count": len(data_rows),
    }


def main() -> int:
    args = parse_args()
    output_path = Path(args.output).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        if args.source == "bundesbank":
            source_url = build_bundesbank_url(args)
        elif args.source == "boe":
            source_url = build_boe_url(args)
        elif args.source == "rba":
            source_url = args.url or RBA_F2_URL
        elif args.source == "boc":
            source_url = build_boc_url(args)
        elif args.source == "boj":
            source_url = args.url or BOJ_SEARCH_URL
        else:
            raise ValueError(f"Unknown source {args.source!r}.")

        payload = fetch_bytes(source_url)
        if args.format == "raw":
            output_path.write_bytes(payload)
            print(f"Saved raw central_banks payload to {output_path}")
            print(f"source={args.source} URL: {source_url}")
            return 0

        if args.source == "bundesbank":
            text = payload.decode("utf-8", errors="replace")
            if args.bundesbank_format == "csv":
                header, rows = parse_csv_text(text)
                body: dict[str, object] = {"format": "csv", "header": header, "rows": rows, "row_count": len(rows)}
            else:
                body = {"format": "sdmx-json", **parse_bundesbank_sdmx(text)}
        elif args.source == "boe":
            header, rows = parse_csv_text(payload.decode("utf-8-sig", errors="replace"))
            body = {"format": "csv", "header": header, "rows": rows, "row_count": len(rows)}
        elif args.source == "rba":
            body = {"format": "xlsx", **parse_rba_xlsx(payload, args.sheet)}
        elif args.source == "boc":
            text = payload.decode("utf-8", errors="replace")
            try:
                body = {"format": "json", "payload": json.loads(text)}
            except json.JSONDecodeError:
                header, rows = parse_csv_text(text)
                body = {"format": "csv", "header": header, "rows": rows, "row_count": len(rows)}
        elif args.source == "boj":
            text = payload.decode("utf-8", errors="replace")
            tables = parse_html_tables(text)
            body = {
                "format": "html",
                "warning": "BoJ Time-Series Search is fragile; prefer FRED IRLTLT01JPM156N for 10Y JGB.",
                "tables": tables,
                "table_count": len(tables),
            }
        else:
            raise ValueError(f"Unknown source {args.source!r}.")

        artifact: dict[str, object] = {
            "provider": "central_banks",
            "sub_source": args.source,
            "source_url": source_url,
            **body,
        }
        output_path.write_text(json.dumps(artifact, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
        print(f"Saved parsed central_banks artifact to {output_path}")
        print(f"sub_source={args.source}")
        return 0
    except (HTTPError, URLError, OSError, RuntimeError, ValueError, json.JSONDecodeError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
