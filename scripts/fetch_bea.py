#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import html
import io
import json
import re
import sys
import zipfile
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urljoin, urlparse
from urllib.request import Request, urlopen


USER_AGENT = "Codex agent-market-data-terminal/1.0"
DATA_JSON_URL = "https://apps.bea.gov/Data.json"
TAG_PATTERN = re.compile(r"<[^>]+>")
TITLE_PATTERN = re.compile(r"<title>(.*?)</title>", re.IGNORECASE | re.DOTALL)
SUPPORTED_FILE_FORMATS = {"xlsx", "xls", "csv", "zip"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch no-auth BEA public files by direct URL, release page, or Data.json catalog.")
    parser.add_argument("--mode", required=True, choices=["direct", "release-page", "catalog"])
    parser.add_argument("--url", help="Direct BEA file URL or release-page URL, depending on mode.")
    parser.add_argument("--catalog-format", choices=["xlsx", "csv", "zip"], help="Optional file-format filter for catalog mode.")
    parser.add_argument("--match", help="Optional case-insensitive substring filter for release-page or catalog discovery.")
    parser.add_argument("--format", required=True, choices=["raw", "parsed"])
    parser.add_argument("--output", required=True, help="Output path.")
    return parser.parse_args()


def validate_args(args: argparse.Namespace) -> None:
    if args.mode in {"direct", "release-page"} and not args.url:
        raise ValueError("--url is required in direct and release-page modes.")
    if args.mode == "catalog" and args.url:
        raise ValueError("--url is not used in catalog mode.")
    if args.mode != "catalog" and args.catalog_format:
        raise ValueError("--catalog-format is only valid in catalog mode.")


def fetch_response(url: str) -> tuple[bytes, dict[str, str]]:
    request = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(request, timeout=90) as response:
        headers = {key.lower(): value for key, value in response.headers.items()}
        return response.read(), headers


def fetch_bytes(url: str) -> bytes:
    raw_bytes, _ = fetch_response(url)
    return raw_bytes


def fetch_text(url: str) -> str:
    return fetch_bytes(url).decode("utf-8", errors="replace")


def clean_text(value: str) -> str:
    text = TAG_PATTERN.sub(" ", value)
    text = html.unescape(text).replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def parse_scalar(value: str) -> object:
    cleaned = value.strip()
    if cleaned == "":
        return None
    if re.fullmatch(r"-?\d+", cleaned):
        return int(cleaned)
    if re.fullmatch(r"-?\d+\.\d+", cleaned):
        return float(cleaned)
    return cleaned


def normalize_cell(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean_text(str(value))


def trim_row(values: list[object]) -> list[object]:
    trimmed = list(values)
    while trimmed and trimmed[-1] in ("", None):
        trimmed.pop()
    return trimmed


def looks_like_header(row: list[object]) -> bool:
    if len(row) < 2:
        return False
    strings = [value for value in row if isinstance(value, str) and value]
    if len(strings) < 2:
        return False
    lowered = [value.lower() for value in strings]
    return len(lowered) == len(set(lowered))


def rows_to_table(rows: list[list[object]]) -> dict[str, object]:
    non_empty = [trim_row(row) for row in rows if any(cell not in ("", None) for cell in row)]
    if not non_empty:
        return {"row_count": 0, "rows": []}

    header = non_empty[0]
    if looks_like_header(header):
        columns = [str(value) if value is not None else "" for value in header]
        data_rows = []
        for row in non_empty[1:]:
            row_values = list(row) + [None] * max(0, len(columns) - len(row))
            data_rows.append({column: row_values[index] if index < len(row_values) else None for index, column in enumerate(columns)})
        return {"columns": columns, "row_count": len(data_rows), "rows": data_rows}

    return {"row_count": len(non_empty), "rows": non_empty}


def parse_csv_bytes(raw_bytes: bytes) -> dict[str, object]:
    text = raw_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = []
    for row in reader:
        normalized_row = []
        for cell in row:
            parsed = parse_scalar(cell)
            normalized_row.append(parsed if isinstance(parsed, (int, float)) else cell.strip())
        rows.append(normalized_row)
    normalized_rows: list[list[object]] = []
    for row in rows:
        normalized_rows.append([None if cell == "" else cell for cell in row])
    table = rows_to_table(normalized_rows)
    table["kind"] = "csv"
    return table


def import_openpyxl():
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("Parsed BEA XLSX support requires the optional 'openpyxl' package.") from exc
    return openpyxl


def parse_workbook_rows(row_iterable) -> list[list[object]]:
    rows: list[list[object]] = []
    for row in row_iterable:
        normalized = [normalize_cell(value) for value in row]
        rows.append(normalized)
    return rows


def parse_xlsx_bytes(raw_bytes: bytes) -> dict[str, object]:
    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    sheets = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = parse_workbook_rows(worksheet.iter_rows(values_only=True))
        sheet_artifact = {"name": sheet_name}
        sheet_artifact.update(rows_to_table(rows))
        sheets.append(sheet_artifact)

    properties = workbook.properties
    return {
        "kind": "xlsx",
        "workbook_properties": {
            "title": properties.title,
            "subject": properties.subject,
            "creator": properties.creator,
            "created": properties.created.isoformat() if getattr(properties, "created", None) else None,
            "modified": properties.modified.isoformat() if getattr(properties, "modified", None) else None,
        },
        "sheets": sheets,
    }


def import_xlrd_optional():
    try:
        import xlrd  # type: ignore
    except ImportError:
        return None
    return xlrd


def parse_xls_bytes(raw_bytes: bytes) -> dict[str, object]:
    xlrd = import_xlrd_optional()
    if xlrd is None:
        raise RuntimeError("Parsed BEA XLS support requires the optional 'xlrd' package.")
    workbook = xlrd.open_workbook(file_contents=raw_bytes)
    sheets = []
    for sheet in workbook.sheets():
        rows = []
        for row_index in range(sheet.nrows):
            rows.append([normalize_cell(value) for value in sheet.row_values(row_index)])
        sheet_artifact = {"name": sheet.name}
        sheet_artifact.update(rows_to_table(rows))
        sheets.append(sheet_artifact)
    return {"kind": "xls", "sheets": sheets}


def file_format_for_path(path_or_url: str) -> str | None:
    suffix = Path(urlparse(path_or_url).path).suffix.lower().lstrip(".")
    return suffix if suffix in SUPPORTED_FILE_FORMATS else None


def parse_zip_bytes(raw_bytes: bytes) -> dict[str, object]:
    files = []
    with zipfile.ZipFile(io.BytesIO(raw_bytes)) as archive:
        for member in archive.infolist():
            if member.is_dir():
                continue
            member_name = member.filename
            entry: dict[str, object] = {
                "name": member_name,
                "size": member.file_size,
                "format": file_format_for_path(member_name),
            }
            if entry["format"] in {"csv", "xlsx", "xls"}:
                member_bytes = archive.read(member)
                try:
                    entry["parsed"] = parse_file_bytes(member_bytes, member_name)
                except RuntimeError as exc:
                    entry["parse_error"] = str(exc)
            files.append(entry)
    return {"kind": "zip", "files": files}


def parse_file_bytes(raw_bytes: bytes, source_url: str) -> dict[str, object]:
    file_format = file_format_for_path(source_url)
    if file_format == "csv":
        return parse_csv_bytes(raw_bytes)
    if file_format == "xlsx":
        return parse_xlsx_bytes(raw_bytes)
    if file_format == "xls":
        return parse_xls_bytes(raw_bytes)
    if file_format == "zip":
        return parse_zip_bytes(raw_bytes)
    raise RuntimeError(f"Unsupported BEA file format for parsed mode: {source_url}")


class LinkCollector(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.links: list[dict[str, str]] = []
        self.current_href: str | None = None
        self.current_text_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag != "a":
            return
        attr_map = {key: value or "" for key, value in attrs}
        href = attr_map.get("href", "")
        if href:
            self.current_href = href
            self.current_text_parts = []

    def handle_data(self, data: str) -> None:
        if self.current_href is not None:
            self.current_text_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag != "a" or self.current_href is None:
            return
        self.links.append({"href": self.current_href, "text": clean_text("".join(self.current_text_parts))})
        self.current_href = None
        self.current_text_parts = []


def extract_title(html_text: str) -> str | None:
    match = TITLE_PATTERN.search(html_text)
    return clean_text(match.group(1)) if match else None


def discover_release_page(url: str, match_text: str | None) -> tuple[str, dict[str, object]]:
    html_text = fetch_text(url)
    parser = LinkCollector()
    parser.feed(html_text)
    parser.close()

    candidates: list[dict[str, str]] = []
    seen: set[str] = set()
    for link in parser.links:
        href = urljoin(url, link["href"])
        file_format = file_format_for_path(href)
        if file_format not in {"xlsx", "xls", "csv", "zip"}:
            continue
        haystack = f"{href} {link['text']}".lower()
        if match_text and match_text.lower() not in haystack:
            continue
        if href in seen:
            continue
        seen.add(href)
        candidates.append({"url": href, "label": link["text"], "format": file_format})

    if not candidates:
        raise RuntimeError("No downloadable BEA file links matched on the release page.")

    return candidates[0]["url"], {
        "release_page_url": url,
        "release_page_title": extract_title(html_text),
        "candidates": candidates,
    }


def candidate_matches(candidate: dict[str, object], match_text: str | None) -> bool:
    if not match_text:
        return True
    haystack_parts = [
        str(candidate.get("dataset_title", "")),
        str(candidate.get("dataset_identifier", "")),
        str(candidate.get("dataset_description", "")),
        str(candidate.get("distribution_description", "")),
        str(candidate.get("url", "")),
    ]
    return match_text.lower() in " ".join(haystack_parts).lower()


def discover_catalog(catalog_format: str | None, match_text: str | None) -> tuple[str, dict[str, object]]:
    catalog_payload = json.loads(fetch_bytes(DATA_JSON_URL).decode("utf-8"))
    datasets = catalog_payload.get("dataset", [])
    if not isinstance(datasets, list):
        raise RuntimeError("BEA Data.json did not include the expected dataset list.")

    candidates: list[dict[str, object]] = []
    for dataset in datasets:
        if not isinstance(dataset, dict):
            continue
        distributions = dataset.get("distribution", [])
        if not isinstance(distributions, list):
            continue
        for distribution in distributions:
            if not isinstance(distribution, dict):
                continue
            url = distribution.get("downloadURL")
            if not url and isinstance(distribution.get("accessURL"), str) and file_format_for_path(str(distribution.get("accessURL"))):
                url = distribution.get("accessURL")
            if not isinstance(url, str):
                continue
            file_format = file_format_for_path(url)
            if file_format not in {"xlsx", "csv", "zip"}:
                continue
            candidate = {
                "url": url,
                "format": file_format,
                "dataset_title": dataset.get("title"),
                "dataset_identifier": dataset.get("identifier"),
                "dataset_description": dataset.get("description"),
                "distribution_description": distribution.get("description"),
            }
            if catalog_format and file_format != catalog_format:
                continue
            if not candidate_matches(candidate, match_text):
                continue
            candidates.append(candidate)

    if not candidates:
        raise RuntimeError("No downloadable BEA catalog distributions matched the requested filters.")

    return str(candidates[0]["url"]), {"catalog_url": DATA_JSON_URL, "candidates": candidates}


def fetch_first_available_candidate(candidates: object, fallback_url: str) -> tuple[str, bytes, dict[str, str], int]:
    if not isinstance(candidates, list) or not candidates:
        raw_bytes, headers = fetch_response(fallback_url)
        return fallback_url, raw_bytes, headers, 0

    last_error: Exception | None = None
    for index, candidate in enumerate(candidates):
        if not isinstance(candidate, dict):
            continue
        url = candidate.get("url")
        if not isinstance(url, str):
            continue
        try:
            raw_bytes, headers = fetch_response(url)
            return url, raw_bytes, headers, index
        except (HTTPError, URLError, OSError) as exc:
            last_error = exc
            continue

    if last_error is not None:
        raise RuntimeError(f"No reachable BEA download candidates matched the requested filters. Last error: {last_error}") from last_error
    raise RuntimeError("No reachable BEA download candidates matched the requested filters.")


def main() -> int:
    args = parse_args()
    validate_args(args)
    output_path = Path(args.output).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        context: dict[str, object] = {"mode": args.mode}
        if args.mode == "direct":
            source_url = str(args.url)
            raw_bytes, headers = fetch_response(source_url)
        elif args.mode == "release-page":
            source_url, discovery = discover_release_page(str(args.url), args.match)
            context.update(discovery)
            source_url, raw_bytes, headers, candidate_index = fetch_first_available_candidate(context.get("candidates", []), source_url)
            context["selected_candidate_index"] = candidate_index
        else:
            source_url, discovery = discover_catalog(args.catalog_format, args.match)
            context.update(discovery)
            source_url, raw_bytes, headers, candidate_index = fetch_first_available_candidate(context.get("candidates", []), source_url)
            context["selected_candidate_index"] = candidate_index

        if args.format == "raw":
            output_path.write_bytes(raw_bytes)
            print(f"Saved raw BEA payload to {output_path}")
            print(f"URL: {source_url}")
            return 0

        parsed = parse_file_bytes(raw_bytes, source_url)
        artifact = {
            "provider": "bea",
            "mode": args.mode,
            "source_url": source_url,
            "file_name": Path(urlparse(source_url).path).name,
            "file_format": file_format_for_path(source_url),
            "content_type": headers.get("content-type"),
            "parsed": parsed,
        }
        artifact.update(context)
        output_path.write_text(json.dumps(artifact, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"Saved parsed BEA artifact to {output_path}")
        return 0
    except (HTTPError, URLError, OSError, RuntimeError, ValueError, zipfile.BadZipFile, json.JSONDecodeError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
